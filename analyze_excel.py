import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('manuais/coletes/IM.022 - Coletes Insufláveis Ficha de Verif. Múltipla.xlsx')
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print("\n=== CONTENT (first 50 rows) ===\n")

# Read all data
for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True), 1):
    row_values = [str(cell) if cell is not None else "" for cell in row]
    # Print non-empty rows
    if any(row_values):
        print(f"Row {i}: {row_values[:12]}")  # Print first 12 columns

print("\n=== ALL SHEETS ===")
print(wb.sheetnames)

# Print more detail on first sheet
print("\n=== DETAILED FIRST 15 ROWS ===")
ws = wb[wb.sheetnames[0]]
for i, row in enumerate(ws.iter_rows(max_row=15, max_col=10, values_only=True), 1):
    print(f"Row {i}: {row}")
