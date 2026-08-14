import openpyxl
from openpyxl import load_workbook
import os

def preencher_template(dados, caminho_template, caminho_saida):
    wb = load_workbook(caminho_template)
    ws = wb.active
    # Exemplo: preencher células específicas
    ws['B2'] = dados.get('marca', '')
    ws['B3'] = dados.get('modelo', '')
    ws['B4'] = dados.get('serial', '')
    ws['B5'] = dados.get('data_fabrico', '')
    ws['B6'] = dados.get('capacidade', '')
    ws['B7'] = dados.get('proprietario', '')
    # ... continue para os outros campos necessários
    wb.save(caminho_saida)

if __name__ == "__main__":
    # Exemplo de uso
    dados = {
        'marca': 'RFD',
        'modelo': 'XPTO',
        'serial': '12345',
        'data_fabrico': '2022-01-01',
        'capacidade': 6,
        'proprietario': 'João'
    }
    caminho_template = os.path.join('templates', 'quadro de inspecao.xltx')
    caminho_saida = os.path.join('templates', 'quadro_de_inspecao_preenchido.xlsx')
    preencher_template(dados, caminho_template, caminho_saida)
