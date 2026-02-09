#!/usr/bin/env python3
"""
Script para analisar todos os certificados Excel, considerando todas as abas.
Extrai o máximo de dados possível de cada aba e salva um relatório CSV.
"""

import openpyxl
import pandas as pd

from pathlib import Path
import csv
import os
import re
from datetime import datetime

cert_dir = Path(r"C:\Users\julio\Desktop\APLICACAO MASTER\LIFERAFT1.0\gestor-naval-pro\OREY DIGITAL 2026\2025\CERTIFICADOS 2025")
output_csv = Path(r"C:\Users\julio\Desktop\APLICACAO MASTER\LIFERAFT1.0\gestor-naval-pro\certificados-aba-relatorio.csv")

fields = ["arquivo", "aba", "linha", "embarcacao", "numeroSerie", "marca", "modelo", "capacidade", "tipoPack", "dataFabrico", "armador", "dataInspecao", "observacao", "raw"]

def normalize(val):
    if not val: return ''
    return str(val).replace('\n', ' ').replace(';', ',').strip()

def extract_date(text):
    if not text: return ''
    # Busca por datas dd/mm/yyyy ou yyyy-mm-dd
    match = re.search(r'(\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2})', text)
    return match.group(0) if match else ''

def extract_number(text):
    if not text: return ''
    match = re.search(r'\d+', text)
    return match.group(0) if match else ''

with open(output_csv, "w", newline='', encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(fields)
    for excel_file in cert_dir.glob("*.xlsx"):
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
                row_str = [normalize(cell) for cell in row]
                row_joined = ' | '.join(row_str)
                # IA heurística: busca por padrões e campos
                embarcacao = next((v for v in row_str if re.search(r'embarca', v, re.I)), '')
                numeroSerie = next((v for v in row_str if re.search(r'serie|serial', v, re.I)), '')
                marca = next((v for v in row_str if re.search(r'marca', v, re.I)), '')
                modelo = next((v for v in row_str if re.search(r'modelo', v, re.I)), '')
                capacidade = next((v for v in row_str if re.search(r'capacidade|lotacao', v, re.I)), '')
                tipoPack = next((v for v in row_str if re.search(r'pack', v, re.I)), '')
                dataFabrico = next((v for v in row_str if re.search(r'fabrico|fabricacao', v, re.I)), '')
                armador = next((v for v in row_str if re.search(r'armador', v, re.I)), '')
                dataInspecao = extract_date(row_joined)
                observacao = next((v for v in row_str if re.search(r'observa|nota|coment', v, re.I)), '')
                # Salva linha se algum campo relevante for encontrado
                if any([embarcacao, numeroSerie, marca, modelo, capacidade, tipoPack, dataFabrico, armador, dataInspecao, observacao]):
                    writer.writerow([
                        excel_file.name, sheet_name, i, embarcacao, numeroSerie, marca, modelo, capacidade, tipoPack, dataFabrico, armador, dataInspecao, observacao, row_joined
                    ])
        wb.close()
print(f"Relatório evoluído salvo em {output_csv}")
