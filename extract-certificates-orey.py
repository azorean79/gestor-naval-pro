#!/usr/bin/env python3
"""
Script para extrair dados dos certificados OREY DIGITAL 2026
Extrai: certificado, data inspeção, próxima inspeção, armador, bandeira,
        série, retenida, data fabrico, pack, cilindros, validades, testes, marcas/modelos
"""

import openpyxl
from pathlib import Path
import json
from datetime import datetime
import re
from typing import Dict, List, Any, Optional

def extract_certificate_number(ws) -> Optional[str]:
    """Extrai número do certificado da linha 3"""
    try:
        for row in ws.iter_rows(min_row=3, max_row=5, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and "AZ" in cell.upper():
                    # Extrair número do certificado (ex: AZ25-001)
                    match = re.search(r'(AZ\d{2}-\d{3})', cell, re.IGNORECASE)
                    if match:
                        return match.group(1).upper()
    except Exception as e:
        print(f"Erro ao extrair número certificado: {e}")
    return None

def parse_date(value: Any) -> Optional[str]:
    """Converte diversos formatos de data para ISO"""
    if value is None:
        return None
    
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    
    if isinstance(value, str):
        # Formatos comuns: MM/YYYY, DD/MM/YYYY, MM-DD-YYYY
        # MM/YYYY
        match = re.match(r'(\d{2})/(\d{4})', value)
        if match:
            month, year = match.groups()
            return f"{year}-{month}-01"
        
        # MM/YY
        match = re.match(r'(\d{2})/(\d{2})', value)
        if match:
            month, year = match.groups()
            full_year = f"20{year}" if int(year) < 50 else f"19{year}"
            return f"{full_year}-{month}-01"
        
        # DD-MM-YYYY ou DD/MM/YYYY
        match = re.match(r'(\d{2})[-/](\d{2})[-/](\d{4})', value)
        if match:
            day, month, year = match.groups()
            return f"{year}-{month}-{day}"
    
    return str(value)

def extract_liferaft_data(ws) -> Dict[str, Any]:
    """Extrai dados principais da jangada da ABA CERTIFICADO"""
    data = {
        'marca_modelo': None,
        'capacidade': None,
        'numero_serie': None,
        'data_fabrico': None,
        'tipo_tela': None,
        'comprimento_retenida': None,
        'cilindro_serie': None,
        'cilindro_co2': None,
        'cilindro_n2': None,
        'cilindro_teste_hidraulico': None,
        'pack_tipo': None,
        'pack_validade': None,
    }
    
    try:
        # Linha 12: Marca/Modelo, Capacidade, Série, Data Fabrico
        row_12 = list(ws.iter_rows(min_row=12, max_row=12, values_only=True))[0]
        row_12_clean = [cell for cell in row_12 if cell is not None]
        
        if len(row_12_clean) >= 3:
            data['marca_modelo'] = str(row_12_clean[0]).strip() if row_12_clean[0] else None
            data['capacidade'] = int(row_12_clean[1]) if row_12_clean[1] else None
            data['numero_serie'] = str(row_12_clean[2]).strip() if row_12_clean[2] else None
            if len(row_12_clean) >= 4:
                data['data_fabrico'] = parse_date(row_12_clean[3])
        
        # Linha 15: Tipo tela, Comprimento retenida
        row_15 = list(ws.iter_rows(min_row=15, max_row=15, values_only=True))[0]
        row_15_clean = [cell for cell in row_15 if cell is not None]
        
        if len(row_15_clean) >= 2:
            data['tipo_tela'] = str(row_15_clean[0]).strip() if row_15_clean[0] else None
            if isinstance(row_15_clean[1], (int, float)):
                data['comprimento_retenida'] = float(row_15_clean[1])
        
        # Linha 19: Dados do cilindro
        row_19 = list(ws.iter_rows(min_row=19, max_row=19, values_only=True))[0]
        row_19_clean = [cell for cell in row_19 if cell is not None]
        
        if len(row_19_clean) >= 3:
            data['cilindro_serie'] = str(row_19_clean[0]).strip() if row_19_clean[0] else None
            data['cilindro_co2'] = float(row_19_clean[1]) if isinstance(row_19_clean[1], (int, float)) else None
            
            # N2 pode estar na posição 2 ou 3
            for i in range(2, min(len(row_19_clean), 5)):
                if isinstance(row_19_clean[i], (int, float)) and row_19_clean[i] < 1:
                    data['cilindro_n2'] = float(row_19_clean[i])
                    break
                elif isinstance(row_19_clean[i], str) and '/' in str(row_19_clean[i]):
                    data['cilindro_teste_hidraulico'] = parse_date(row_19_clean[i])
                    break
        
        # Linha 26: Pack tipo e validade
        row_26 = list(ws.iter_rows(min_row=26, max_row=26, values_only=True))[0]
        row_26_clean = [cell for cell in row_26 if cell is not None]
        
        if len(row_26_clean) >= 2:
            # Tipo do pack (ex: SOLAS B, R, SIMPL. REDUZ.)
            data['pack_tipo'] = str(row_26_clean[1]).strip() if row_26_clean[1] else None
            # Validade do pack
            if len(row_26_clean) >= 4:
                data['pack_validade'] = parse_date(row_26_clean[3])
        
    except Exception as e:
        print(f"Erro ao extrair dados da jangada: {e}")
    
    return data

def extract_ship_data(ws_quadro) -> Dict[str, Any]:
    """Extrai dados do navio da ABA QUADRO"""
    data = {
        'navio_nome': None,
        'armador': None,
        'bandeira': None,
    }
    
    try:
        # Linha 7: Série, Nome do Navio, Marca/Modelo
        row_7 = list(ws_quadro.iter_rows(min_row=7, max_row=7, values_only=True))[0]
        row_7_clean = [cell for cell in row_7 if cell is not None]
        
        if len(row_7_clean) >= 4:
            data['navio_nome'] = str(row_7_clean[3]).strip() if row_7_clean[3] else None
        
        # Procurar armador e bandeira nas próximas linhas (geralmente não estão no certificado padrão)
        # Vamos deixar como None por enquanto, mas preparado para expansão
        
    except Exception as e:
        print(f"Erro ao extrair dados do navio: {e}")
    
    return data

def extract_component_validities(ws_quadro) -> List[Dict[str, Any]]:
    """Extrai validades de componentes da ABA QUADRO"""
    components = []
    
    try:
        # Componentes começam na linha 10-80
        for row in ws_quadro.iter_rows(min_row=10, max_row=80, values_only=True):
            row_clean = [cell for cell in row if cell is not None and str(cell).strip()]
            
            if not row_clean:
                continue
            
            # Procurar datas de validade (formato MM/YYYY ou datetime)
            for i, cell in enumerate(row_clean):
                validade = None
                componente = None
                
                # Se for datetime
                if isinstance(cell, datetime):
                    validade = cell.strftime('%Y-%m-%d')
                    # Componente é a célula anterior
                    if i > 0:
                        componente = str(row_clean[i-1]).strip()
                
                # Se for string com data
                elif isinstance(cell, str):
                    # Formato MM/YYYY ou MM/YY
                    match = re.match(r'(\d{2})/(\d{2,4})', cell)
                    if match:
                        month, year = match.groups()
                        if len(year) == 2:
                            year = f"20{year}" if int(year) < 50 else f"19{year}"
                        validade = f"{year}-{month}-01"
                        # Componente é a célula anterior
                        if i > 0:
                            componente = str(row_clean[i-1]).strip()
                
                if validade and componente:
                    components.append({
                        'componente': componente,
                        'validade': validade
                    })
    
    except Exception as e:
        print(f"Erro ao extrair validades de componentes: {e}")
    
    return components

def extract_inspection_dates(ws) -> Dict[str, Optional[str]]:
    """Extrai datas de inspeção do certificado"""
    dates = {
        'data_inspecao': None,
        'proxima_inspecao': None,
    }
    
    try:
        # Datas geralmente aparecem no final do certificado
        # Procurar em linhas 60-77
        for row in ws.iter_rows(min_row=60, max_row=77, values_only=True):
            row_clean = [cell for cell in row if cell is not None]
            
            for i, cell in enumerate(row_clean):
                cell_str = str(cell).strip() if cell else ""
                
                # Procurar padrões de data
                if isinstance(cell, datetime):
                    # Se ainda não temos data de inspeção, assumir que é esta
                    if not dates['data_inspecao']:
                        dates['data_inspecao'] = cell.strftime('%Y-%m-%d')
                    elif not dates['proxima_inspecao']:
                        dates['proxima_inspecao'] = cell.strftime('%Y-%m-%d')
                
                # Procurar texto indicativo
                if "DATE OF INSPECTION" in cell_str.upper() or "DATA DE INSPEÇÃO" in cell_str.upper():
                    if i + 1 < len(row_clean):
                        dates['data_inspecao'] = parse_date(row_clean[i + 1])
                
                if "PRÓXIMA" in cell_str.upper() or "NEXT" in cell_str.upper():
                    if i + 1 < len(row_clean):
                        dates['proxima_inspecao'] = parse_date(row_clean[i + 1])
    
    except Exception as e:
        print(f"Erro ao extrair datas de inspeção: {e}")
    
    return dates

def extract_tests(ws_quadro) -> List[str]:
    """Extrai testes realizados da jangada"""
    tests = []
    
    try:
        # Procurar palavras-chave de testes nas últimas linhas
        keywords = [
            'TESTE', 'TEST', 'PRESSURE', 'LEAK', 'INFLATION', 
            'INSUFLA', 'PRESSÃO', 'FUGA', 'VÁLVULA'
        ]
        
        for row in ws_quadro.iter_rows(min_row=40, max_row=84, values_only=True):
            row_str = ' '.join([str(cell) for cell in row if cell])
            
            for keyword in keywords:
                if keyword in row_str.upper():
                    tests.append(row_str.strip())
                    break
    
    except Exception as e:
        print(f"Erro ao extrair testes: {e}")
    
    return tests

def process_certificate(excel_path: Path) -> Optional[Dict[str, Any]]:
    """Processa um ficheiro de certificado Excel completo"""
    print(f"\nProcessando: {excel_path.name}")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Verificar se tem as duas abas esperadas
        if 'CERTIFICADO' not in wb.sheetnames or 'QUADRO' not in wb.sheetnames:
            print(f"  [AVISO]: Abas esperadas nao encontradas. Abas disponiveis: {wb.sheetnames}")
            wb.close()
            return None
        
        ws_cert = wb['CERTIFICADO']
        ws_quadro = wb['QUADRO']
        
        # Extrair todos os dados
        cert_data = {
            'ficheiro': excel_path.name,
            'numero_certificado': extract_certificate_number(ws_cert),
        }
        
        # Dados da jangada
        liferaft_data = extract_liferaft_data(ws_cert)
        cert_data.update(liferaft_data)
        
        # Dados do navio
        ship_data = extract_ship_data(ws_quadro)
        cert_data.update(ship_data)
        
        # Datas de inspeção
        dates = extract_inspection_dates(ws_cert)
        cert_data.update(dates)
        
        # Componentes e validades
        cert_data['componentes'] = extract_component_validities(ws_quadro)
        
        # Testes
        cert_data['testes'] = extract_tests(ws_quadro)
        
        wb.close()
        
        print(f"  [OK] Certificado: {cert_data['numero_certificado']}")
        print(f"       Serie: {cert_data['numero_serie']}")
        print(f"       Navio: {cert_data['navio_nome']}")
        print(f"       Marca/Modelo: {cert_data['marca_modelo']}")
        print(f"       Componentes extraidos: {len(cert_data['componentes'])}")
        
        return cert_data
        
    except Exception as e:
        print(f"  [ERRO]: {str(e)}")
        return None

def extract_brand_model(marca_modelo_str: str) -> tuple[Optional[str], Optional[str]]:
    """Separa marca e modelo de uma string"""  
    if not marca_modelo_str:
        return None, None
    
    # Padrões conhecidos
    patterns = [
        (r'^(RFD)\s+(.+)$', lambda m: (m.group(1), m.group(2))),
        (r'^(DSB)\s+(.+)$', lambda m: (m.group(1), m.group(2))),
        (r'^(VIKING)\s+(.+)$', lambda m: (m.group(1), m.group(2))),
        (r'^(EUROVINIL)\s+(.+)$', lambda m: (m.group(1), m.group(2))),
    ]
    
    for pattern, extractor in patterns:
        match = re.match(pattern, marca_modelo_str, re.IGNORECASE)
        if match:
            return extractor(match)
    
    # Se não corresponder a nenhum padrão, tentar dividir pela primeira palavra
    parts = marca_modelo_str.split(None, 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    
    return marca_modelo_str, None

def main():
    """Função principal"""
    cert_dir = Path(r"C:\Users\julio\Desktop\APLICACAO MASTER\LIFERAFT1.0\gestor-naval-pro\OREY DIGITAL 2026\2025\CERTIFICADOS 2025")
    
    if not cert_dir.exists():
        print(f"[ERRO]: Diretorio nao encontrado: {cert_dir}")
        return
    
    # Listar todos os ficheiros Excel
    excel_files = sorted(cert_dir.glob("*.xlsx"))
    print(f"\n{'='*80}")
    print(f"EXTRACAO DE CERTIFICADOS OREY DIGITAL 2026")
    print(f"{'='*80}")
    print(f"Encontrados {len(excel_files)} certificados para processar\n")
    
    # Processar todos os certificados
    all_certificates = []
    brands_models = set()
    
    for excel_file in excel_files:
        cert_data = process_certificate(excel_file)
        if cert_data:
            all_certificates.append(cert_data)
            
            # Extrair marcas e modelos únicos
            if cert_data.get('marca_modelo'):
                marca, modelo = extract_brand_model(cert_data['marca_modelo'])
                if marca and modelo:
                    brands_models.add((marca, modelo))
    
    # Guardar resultados
    output_file = Path("certificados-orey-2025-extracted.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_certificates, f, ensure_ascii=False, indent=2)
    
    print(f"\n{'='*80}")
    print(f"RESUMO DA EXTRACAO")
    print(f"{'='*80}")
    print(f"Total de certificados processados: {len(all_certificates)}")
    print(f"Total de marcas/modelos unicos: {len(brands_models)}")
    print(f"\nMarcas e Modelos encontrados:")
    for marca, modelo in sorted(brands_models):
        print(f"  - {marca} - {modelo}")
    
    print(f"\n[OK] Dados extraidos guardados em: {output_file.absolute()}")
    
    # Estatisticas adicionais
    total_components = sum(len(cert.get('componentes', [])) for cert in all_certificates)
    certs_with_ship = sum(1 for cert in all_certificates if cert.get('navio_nome'))
    certs_with_dates = sum(1 for cert in all_certificates if cert.get('data_inspecao'))
    
    print(f"\nEstatisticas:")
    print(f"  - Total de componentes extraidos: {total_components}")
    print(f"  - Certificados com nome de navio: {certs_with_ship}")
    print(f"  - Certificados com data de inspecao: {certs_with_dates}")

if __name__ == "__main__":
    main()
