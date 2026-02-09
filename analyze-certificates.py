#!/usr/bin/env python3
"""
Script para analisar certificados Excel da OREY DIGITAL 2026
Extrai dados de inspeção, componentes, marcas e modelos
"""

import openpyxl
import pandas as pd
from pathlib import Path
import json
from datetime import datetime
import re

def analyze_excel_structure(excel_path):
    """Analisa a estrutura de um ficheiro Excel de certificado"""
    print(f"\n{'='*80}")
    print(f"Analisando: {excel_path.name}")
    print(f"{'='*80}\n")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        print(f"Número de abas: {len(wb.sheetnames)}")
        print(f"Nomes das abas: {wb.sheetnames}\n")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- ABA: {sheet_name} ---")
            print(f"Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")
            
            # Mostrar primeiras 30 linhas com conteúdo
            print("\nPrimeiras linhas:")
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
                # Filtrar células vazias
                row_data = [cell for cell in row if cell is not None and str(cell).strip()]
                if row_data:
                    print(f"Linha {i:2d}: {row_data}")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"ERRO ao analisar {excel_path.name}: {str(e)}")
        return False

def main():
    """Função principal"""
    cert_dir = Path(r"C:\Users\julio\Desktop\APLICACAO MASTER\LIFERAFT1.0\gestor-naval-pro\OREY DIGITAL 2026\2025\CERTIFICADOS 2025")
    
    if not cert_dir.exists():
        print(f"ERRO: Diretório não encontrado: {cert_dir}")
        return
    
    # Listar todos os ficheiros Excel
    excel_files = list(cert_dir.glob("*.xlsx"))
    print(f"\nEncontrados {len(excel_files)} ficheiros Excel")
    
    # Analisar os primeiros 3 ficheiros para entender a estrutura
    print("\n" + "="*80)
    print("ANALISANDO ESTRUTURA DOS CERTIFICADOS (primeiros 3 ficheiros)")
    print("="*80)
    
    for excel_file in excel_files[:3]:
        analyze_excel_structure(excel_file)
        print("\n" + "="*80 + "\n")

if __name__ == "__main__":
    main()
