import sys, os, subprocess, json
sys.path.insert(0, 'D:\\Acores')

node_code = '''
const { PrismaClient } = require("@prisma/client");
const p = new PrismaClient();
async function main() {
  const jangadas = await p.jangada.findMany({
    include: { navio: { select: { nome: true, matricula: true, cliente: { select: { nome: true, nif: true } } } } },
    orderBy: { id: "asc" }
  });
  const data = jangadas.map(j => ({
    id: j.id, serial: j.serial, brand: j.brand, model: j.model, capacity: j.capacity,
    packType: j.packType, containerModel: j.containerModel, launchType: j.launchType,
    dataFabrico: j.dataFabrico, dataInspecao: j.dataInspecao, dataProxInspecao: j.dataProxInspecao,
    ultimoCertificadoNumero: j.ultimoCertificadoNumero, owner: j.owner,
    navioNome: j.navio?.nome || j.shipNameManual || "",
    navioMatricula: j.navio?.matricula || "",
    clienteNome: j.navio?.cliente?.nome || "",
    clienteNif: j.navio?.cliente?.nif || "",
  }));
  console.log(JSON.stringify(data));
}
main().finally(() => { p.$disconnect().catch(() => {}).then(() => process.exit(0)); });
'''

result = subprocess.run(['node', '-e', node_code], capture_output=True, text=True, timeout=60, cwd='D:\\Acores')
jangadas = json.loads(result.stdout)
print(f'Total jangadas: {len(jangadas)}')

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Jangadas Acores"

headers = ['ID', 'Serial', 'Marca', 'Modelo', 'Capacidade', 'Pack Type', 'Contentor',
           'Lancamento', 'Data Fabrico', 'Ultima Inspecao', 'Prox Inspecao',
           'Certificado', 'Proprietario', 'Navio', 'Matricula', 'Cliente', 'Cliente NIF']

header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1A3C6E', end_color='1A3C6E', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

for row, j in enumerate(jangadas, 2):
    values = [j['id'], j['serial'], j['brand'], j['model'], j['capacity'], j['packType'],
              j['containerModel'], j['launchType'], j['dataFabrico'], j['dataInspecao'],
              j['dataProxInspecao'], j['ultimoCertificadoNumero'], j['owner'],
              j['navioNome'], j['navioMatricula'], j['clienteNome'], j['clienteNif']]
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = Font(name='Calibri', size=9)
        cell.border = thin_border

for col in ws.columns:
    max_len = 0
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

ws.freeze_panes = 'A2'

output = 'D:\\Acores\\public\\relatorio_jangadas.xlsx'
wb.save(output)
print(f'Relatorio salvo: {output}')
print(f'Total: {len(jangadas)} jangadas')
