import json, openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

with open('D:\\Acores\\temp_stock.json', encoding='utf-8') as f:
    items = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Catalogo Stock"

headers = ['Referencia', 'Descricao', 'Categoria', 'Stock', 'Stock Minimo',
           'Preco Venda', 'Cod. Fabricante', 'Foto']

hfont = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
hfill = PatternFill(start_color='1A3C6E', end_color='1A3C6E', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hfont; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = border

for row, item in enumerate(items, 2):
    vals = [item['ref'], item['desc'], item['cat'], item['qtd'], item['min'],
            item['preco'], item['fab'], item['foto']]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = Font(name='Calibri', size=9)
        cell.border = border
        if col == 4: cell.alignment = Alignment(horizontal='center')

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 35
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

wb.save('D:\\Acores\\public\\catalogo_stock.xlsx')
print(f'OK: {len(items)} artigos -> public/catalogo_stock.xlsx')