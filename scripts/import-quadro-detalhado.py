import openpyxl
import os
import json
from datetime import datetime
from dotenv import load_dotenv

# Caminho da pasta dos certificados
cert_dir = r"C:\Users\julio\Desktop\APLICACAO MASTER\LIFERAFT1.0\gestor-naval-pro\OREY DIGITAL 2026\2025\CERTIFICADOS 2025"

# Lista de campos a extrair
COMPONENTES = [
    'Facho de mão', 'Facho paraquedas', 'Pote de fumo', 'Água', 'Ração', 'Farmácia', 'Comprimidos',
    'Luz', 'Bateria', 'EPIRB', 'HRU', 'Radar', 'Kit primeiros socorros', 'Cilindro', 'Válvula', 'Cabeça de disparo'
]

# Resultado final
result = []

for fname in os.listdir(cert_dir):
    if not fname.endswith('.xlsx'): continue
    wb = openpyxl.load_workbook(os.path.join(cert_dir, fname), data_only=True)
    if 'QUADRO' not in wb.sheetnames: continue
    quadro = wb['QUADRO']
    certificado = wb['CERTIFICADO']
    dados = {'arquivo': fname, 'componentes': [], 'cilindro': {}, 'valvulas': [], 'luzes': [], 'baterias': []}

    # Extrair componentes e validades
    for row in range(1, quadro.max_row+1):
        for col in range(1, quadro.max_column+1):
            val = quadro.cell(row, col).value
            if val:
                for comp in COMPONENTES:
                    if comp.lower() in str(val).lower():
                        validade = quadro.cell(row, col+1).value
                        dados['componentes'].append({'nome': comp, 'validade': validade})
                if 'cilindro' in str(val).lower():
                    dados['cilindro']['descricao'] = val
                    dados['cilindro']['tara'] = quadro.cell(row, col+1).value
                    dados['cilindro']['peso_bruto'] = quadro.cell(row, col+2).value
                if 'valvula' in str(val).lower():
                    dados['valvulas'].append(val)
                if 'luz' in str(val).lower():
                    validade = quadro.cell(row, col+1).value
                    dados['luzes'].append({'nome': val, 'validade': validade})
                if 'bateria' in str(val).lower():
                    validade = quadro.cell(row, col+1).value
                    dados['baterias'].append({'nome': val, 'validade': validade})
    result.append(dados)

# Salvar resultado
with open('quadro-detalhado-extract.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"Extração concluída. {len(result)} arquivos processados.")