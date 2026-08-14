import json, openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

with open('D:\\Acores\\temp_jangadas.json', encoding='utf-8') as f:
    jangadas = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Jangadas Acores"

headers = ['ID', 'Serial', 'Marca', 'Modelo', 'Capacidade', 'Pack Type', 'Contentor',
           'Lancamento', 'Data Fabrico', 'Ultima Inspecao', 'Prox Inspecao',
           'Certificado', 'Proprietario', 'Navio', 'Matricula', 'Cliente', 'Cliente NIF']

hfont = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
hfill = PatternFill(start_color='1A3C6E', end_color='1A3C6E', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hfont; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = border

for row, j in enumerate(jangadas, 2):
    vals = [j['id'], j['serial'], j['brand'], j['model'], j['capacity'], j['packType'],
            j['containerModel'], j['launchType'], j['dataFabrico'], j['dataInspecao'],
            j['dataProxInspecao'], j['ultimoCertificadoNumero'], j['owner'],
            j['navioNome'], j['navioMatricula'], j['clienteNome'], j['clienteNif']]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = Font(name='Calibri', size=9)
        cell.border = border

for col in ws.columns:
    max_len = max((len(str(c.value)) if c.value else 0) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

ws.freeze_panes = 'A2'
wb.save('D:\\Acores\\public\\relatorio_jangadas.xlsx')
print(f'OK: {len(jangadas)} jangadas -> public/relatorio_jangadas.xlsx')