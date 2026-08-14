import sys, json, os, openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def export_table(table_name, rows, headers, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_name

    hfont = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='1A3C6E', end_color='1A3C6E', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = border

    for row_i, row_data in enumerate(rows, 2):
        for col_i, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font = Font(name='Calibri', size=9); cell.border = border

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 10) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)
    print(f'  {filename} ({len(rows)} registos)')

def create_template(headers, filename, instructions_row):
    wb = openpyxl.Workbook()
    ws = wb.active

    hfont = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='1A3C6E', end_color='1A3C6E', fill_type='solid')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill

    for col, note in enumerate(instructions_row, 1):
        cell = ws.cell(row=2, column=col, value=note)
        cell.font = Font(name='Calibri', size=8, italic=True, color='64748B')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.cell(row=3, column=1).value = '(exemplo)'
    wb.save(filename)
    print(f'  {filename} (template)')

output_dir = 'D:\\Acores\\public\\import_export'
os.makedirs(output_dir, exist_ok=True)

# --- EXPORT ---
print('A exportar dados...')

# Stock
try:
    import subprocess
    r = subprocess.run(['node', '-e', '''
        const{PrismaClient}=require("@prisma/client");
        const p=new PrismaClient();
        async function main(){
          const items=await p.stock.findMany({orderBy:{referencia:"asc"}});
          const d=items.map(s=>[s.referencia||"",s.descricao,s.categoria||"",s.quantidade,s.quantidadeMinima||0,s.precoVenda,s.codigoFabricante||""]);
          console.log(JSON.stringify(d));
        }
        main().finally(()=>{p.$disconnect().catch(()=>{}).then(()=>process.exit(0));});
    '''], capture_output=True, text=True, timeout=30, cwd='D:\\Acores')
    stock_rows = json.loads(r.stdout)
    export_table('Stock', stock_rows, ['Referencia','Descricao','Categoria','Quantidade','Qt Minima','Preco','Cod. Fabricante'],
                 os.path.join(output_dir, 'stock_export.xlsx'))
except: print('  Erro stock')

# Navios
try:
    r = subprocess.run(['node', '-e', '''
        const{PrismaClient}=require("@prisma/client");
        const p=new PrismaClient();
        async function main(){
          const items=await p.navio.findMany({include:{cliente:{select:{nome:true}}},orderBy:{nome:"asc"}});
          const d=items.map(n=>[n.nome,n.matricula,n.ilha||"",n.tipoPesca,n.cliente?.nome||"",n.clienteId||""]);
          console.log(JSON.stringify(d));
        }
        main().finally(()=>{p.$disconnect().catch(()=>{}).then(()=>process.exit(0));});
    '''], capture_output=True, text=True, timeout=30, cwd='D:\\Acores')
    navio_rows = json.loads(r.stdout)
    export_table('Navios', navio_rows, ['Nome','Matricula','Ilha','Tipo Pesca','Cliente','Cliente ID'],
                 os.path.join(output_dir, 'navios_export.xlsx'))
except: print('  Erro navios')

# Clientes
try:
    r = subprocess.run(['node', '-e', '''
        const{PrismaClient}=require("@prisma/client");
        const p=new PrismaClient();
        async function main(){
          const items=await p.cliente.findMany({orderBy:{nome:"asc"}});
          const d=items.map(c=>[c.nome,c.numeroCliente||"",c.nif||"",c.email||"",c.telefone||"",c.telmovel||"",c.morada||"",c.ilha||""]);
          console.log(JSON.stringify(d));
        }
        main().finally(()=>{p.$disconnect().catch(()=>{}).then(()=>process.exit(0));});
    '''], capture_output=True, text=True, timeout=30, cwd='D:\\Acores')
    cliente_rows = json.loads(r.stdout)
    export_table('Clientes', cliente_rows, ['Nome','N. Cliente','NIF','Email','Telefone','Telemovel','Morada','Ilha'],
                 os.path.join(output_dir, 'clientes_export.xlsx'))
except Exception as e: print(f'  Erro clientes: {e}')

# Jangadas
try:
    r = subprocess.run(['node', '-e', '''
        const{PrismaClient}=require("@prisma/client");
        const p=new PrismaClient();
        async function main(){
          const items=await p.jangada.findMany({orderBy:{id:"asc"},take:100});
          const d=items.map(j=>[j.id,j.serial,j.brand||"",j.model||"",j.capacity,j.packType,j.containerModel||"",j.owner||""]);
          console.log(JSON.stringify(d));
        }
        main().finally(()=>{p.$disconnect().catch(()=>{}).then(()=>process.exit(0));});
    '''], capture_output=True, text=True, timeout=30, cwd='D:\\Acores')
    jangada_rows = json.loads(r.stdout)
    export_table('Jangadas', jangada_rows, ['ID','Serial','Marca','Modelo','Capacidade','Pack Type','Contentor','Proprietario'],
                 os.path.join(output_dir, 'jangadas_export.xlsx'))
except: print('  Erro jangadas')

# --- TEMPLATES (vazios para importacao) ---
print('\nA criar templates de importacao...')
create_template(['Nome','NIF','Email','Telefone','Telemovel','Morada','Ilha','Modo Pagamento'],
    os.path.join(output_dir, 'template_importar_clientes.xlsx'),
    ['Nome completo','9 digitos','email@exemplo.pt','+351...','+351...','Rua, n, cod-postal, localidade','Ilha dos Acores','Ex: Numerario, Transferencia'])

create_template(['Nome Navio','Matricula','Ilha','Tipo Pesca','Cliente Nome','Cliente NIF'],
    os.path.join(output_dir, 'template_importar_navios.xlsx'),
    ['Nome do navio','Numero matricula','Ilha de registo','Tipo de pesca','Nome do cliente (existente)','NIF do cliente'])

create_template(['Referencia','Descricao','Categoria','Quantidade','Preco Venda','Cod. Fabricante'],
    os.path.join(output_dir, 'template_importar_stock.xlsx'),
    ['Ref. unica','Descricao completa','Ex: CILINDROS,PIROTECNICOS...','Qtd atual','Preco em euros','Codigo do fabricante'])

create_template(['Serial','Marca','Modelo','Capacidade','Pack Type','Container','Data Fabrico','Proprietario'],
    os.path.join(output_dir, 'template_importar_jangadas.xlsx'),
    ['Serial unico','Marca','Modelo','Numero pessoas','SOLAS A/B, COASTAL...','Modelo contentor','AAAA-MM','Nome armador'])

print(f'\nFicheiros em: {output_dir}')
for f in os.listdir(output_dir):
    size = os.path.getsize(os.path.join(output_dir, f))
    print(f'  {f} ({size/1024:.1f} KB)')